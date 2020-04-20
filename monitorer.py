import  win32gui
import psutil, win32process

import logging

from  openpyxl import load_workbook , Workbook

from pandas import DataFrame, to_datetime
# Load the Pandas libraries with alias 'pd' 

import json
import time
import os
import sys
from pynput.keyboard import Listener

import datetime

from win10toast import ToastNotifier

from infi.systray import SysTrayIcon


#tells if mouse is still
class IdleClassifier(object):

    lastpos = (0,0)    
    lastkeytime = time.time()
    idlethreshold = 10 #seconds  - time in same spot to consider it idle

    #if any key was pressed, refresh tiemmer
    def UpdateKeyTime(self,key):
        self.lastkeytime=time.time()

    #whenever this callback is called, update mouse pos
    def UpdateLastMouse(self):
        try:
            _,_,self.lastpos = win32gui.GetCursorInfo()
        except:
            logging.warning("Mouse Was not acessible at this point")
           


    def Classify(self):
        '''
        Classifies whether user is here or not/ not giving inputs
        '''

        try:
            _, _, (x,y) = win32gui.GetCursorInfo()
        except:
            logging.warning("On Classify. Mouse Was not acessible at this point")
            return "-"
        

        #this does it lol, pos is != and no key was touches in the last 10 seconds
        idle = self.lastpos==(x,y) and time.time()-self.lastkeytime> self.idlethreshold

        self.lastpos=(x,y)

        if idle:
            return "IDLE"
        else:
            return "AWAKE"


def GetWithinTimes(df,start_date,end_date):
        #set time range mask
    mask = (df['Time'] >start_date ) & (df['Time'] <= end_date )

    #get relevant events
    df1 = df.loc[mask]



def ClassClassifier(strings, classfile):
    '''
    Tells which category of thing are you currently doing
    Literraally finds first correspondence in the the info.json and returns it
    '''

    activeclass="Uncategorized"
    detected=False

    #for all classes
    for key in classfile:

        #exit if found
        if detected:
            break
        #chkec if there is key in window name
        for lol in classfile[key]:

            
            #exit if found
            if detected:
                break

            for s in strings:
                #if there is exit
                if lol in s.lower():
                    activeclass=key
                    detected=True
                    break

    return activeclass

 

def GetWindowName():
    #yup
    return win32gui.GetWindowText(win32gui.GetForegroundWindow())

def WriteExcel(ws,index,arr):

    if index<1:
        logging.error("Could not write line, index {0} is invalid,".format(index))
        return

    #array into excel row
    for idx, val in enumerate(arr, start=1):
        ws.cell(row=index, column=idx).value = val

def FindFirstEmptyRow(ws,limit=10000):
    #goes through excel and finds first empty row, (uses column 3 which always has the time value)
    idx=1
    while idx<limit or limit==-1: 
        if ws.cell(row=idx, column=3).value is None:
            return idx

        idx=idx+1

    return -1


def DisplayDailyStats(ws1):
    '''
    Displays the category statistics for the day
    '''

    #Calculate them here:
    data = ws1.values
    # Get the first line in file as a header line
    columns = next(data)[0:]
    # Create a DataFrame based on the second and subsequent lines of data
    df = DataFrame(data, columns=columns)

    #make sure time is in datetime
    df['Time'] = to_datetime(df['Time'],dayfirst=True)  


    #set time range for the current day
    start_date= datetime.datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
    end_date = start_date+datetime.timedelta(days=1) 

    GetWithinTimes(df,start_date,end_date)

    #set time range mask
    mask = (df['Time'] >start_date ) & (df['Time'] <= end_date )

    #get relevant events
    df1 = df.loc[mask]

    #create the time interval between each intervla and the next
    df1['Interval'] = df1['Time'].shift(-1) - df1['Time']

    #count ocurrences for each categorey
    stats = df1.groupby(['Category'])['Interval']

    strvals=""

    for idx,val in stats.sum().items():
        strvals = strvals + idx  + " "+  str(val)[-8:]+"\n"

    print(strvals)
    try:
        ToastNotifier().show_toast("Daily Stats",strvals,icon_path="eye.ico",duration=10,threaded=True)
    except:
        logging.warning("Issue displaying desktop notification")

#state variables
class MyClass:
  isRunning = True

def Stop(state):
    state.isRunning=False

def active_window_process_process():
    pid = win32process.GetWindowThreadProcessId(win32gui.GetForegroundWindow()) #This produces a list of PIDs active window relates to

    try:
        return  psutil.Process(pid[-1]).exe() #pid[-1] is the most likely to survive last longer #.name() if you want hust the file not the path
    except:
        logging.warning("Invalid Pid")
        return "-"



def Monitor(interval,classesfile='info.json',directory="Outputs/"):
    '''
    Where all the magic happens
    '''

    #loads classes and words in it
    with open(classesfile, 'r') as f:
        classes = json.load(f)


    #sets path for file
    filepath = directory + str(datetime.date.today().strftime("%m-%Y")) + ".xlsx"
    
    #filds on top of the excel document
    fields = ['Category','WindowName','Time','MouseState','Path']

    #createæxcel if it does exist else just load it 
    if not os.path.isfile(filepath):  # False
        wb = Workbook()
        wb.save(filepath)

        logging.info("Created New File")
    else:
        try:
            wb = load_workbook(filepath)
            logging.debug("Loaded Existing File")
        except:
            logging.warning("Output file has been corrupted, creating a new one")

            import shutil

            shutil.copyfile(filepath, filepath+str(time.time))
        
            wb = Workbook()
            wb.save(filepath)
            

        #notificationthread.raise_exception()
        


    #fetch first sheet
    ws1 = wb.active
    ws1.title = "Log"
    WriteExcel(ws1,1,fields)
    
    #find first row to put values (search forever)
    count = FindFirstEmptyRow(ws1,limit=-1)
    
    #initialize global variables
    statevars = MyClass()


    #Set Up System Tray Icon

    #Show stats button
    menu_options = (("Show Stats", None, lambda s :DisplayDailyStats(ws1)),)
    #choose icon and how to quit
    systray = SysTrayIcon("eye.ico", "Monitorer", menu_options, on_quit=lambda s :Stop(statevars))
    systray.start()

    #initialize activity classifier
    mouseclassifier = IdleClassifier()

    #open listener to read incoming inputs
    with Listener(on_press=mouseclassifier.UpdateKeyTime) as listener:

        #while not quit        
        while(statevars.isRunning):

            #window name
            string = GetWindowName()

            #date
            dt_string = datetime.datetime.now().strftime("%d/%m/%Y %H:%M:%S")

            

            userstate = mouseclassifier.Classify()

            path = active_window_process_process()

            #category
            curclass = ClassClassifier([string,path], classes)

            output = [ curclass, string,dt_string,userstate,path]

            #write in row
            WriteExcel(ws1,count,output)
            
            #print(output) 
            count = count + 1 

            #maybe not do this here, meh we'll see
            try:
                wb.save(filepath)
            except:
                logging.info("Could not save at this time, excel file is probably open")

            time.sleep(interval)

    saved=False

    while saved == False:
        
        try:
            wb.save(filepath)
            saved=True
        except:
            logging.warn("Cannot Save at this point, retrying in 60 seconds")
            time.sleep(60)

            try:
                toaster = ToastNotifier()
                toaster.show_toast("Work Monitoring",    "Cannot Save File! Please close the excel file bro",    icon_path="eye.ico",    duration=5,    threaded=False)
            except:
                logging.warn("Issue displaying desktop notification")
                
        

    #listener.join()
    wb.save(filepath)
    print("Exiting")

            

def main():
    
    dirName="Logs"
    try:
    # Create target Directory
        os.mkdir(dirName)
        print("Directory " , dirName ,  " Created ") 
    except FileExistsError:
        print("Directory " , dirName ,  " already exists")

    
   
    #logging
    logging.basicConfig(filename="./Logs/"+datetime.datetime.now().strftime("%H-%M-%S %Y-%m-%d") + '.log',level=logging.DEBUG,format='%(asctime)s %(name)-12s %(levelname)-8s %(message)s')
    
    # define a Handler which writes INFO messages or higher to the sys.stderr
    console = logging.StreamHandler()
    console.setLevel(logging.INFO)
    # set a format which is simpler for console use
    formatter = logging.Formatter('%(asctime)s %(name)-12s: %(levelname)-8s %(message)s')
    # tell the handler to use this format
    console.setFormatter(formatter)
    # add the handler to the root logger
    logging.getLogger('').addHandler(console)


    dirName='Outputs'
    try:
    # Create target Directory
        os.mkdir(dirName)
        logging.info("Directory {0} Created".format(dirName) )
        
    except FileExistsError:
        logging.info("Directory {0} already exists".format(dirName) )


    try:
        toaster = ToastNotifier()
        toaster.show_toast("Work Monitoring",    "Monitor Process has been started",    icon_path="eye.ico",    duration=5,    threaded=False)
    except:
        logging.warning("Issue displaying desktop notification")

    # Wait for threaded notification to finish
    #while toaster.notification_active():
    #    time.sleep(0.1)
    

    try:
        Monitor(interval=10)
    except:
        logging.exception("Fatal Error")

if __name__ == "__main__":
    main()